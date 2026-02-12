import asyncio
import logging
import os
import shutil
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from typing import Optional, List, Tuple, Dict
from uuid import uuid4
import datetime
import aiohttp
import pandas as pd
import pyodbc
import requests
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from aiohttp import ClientTimeout
from aiohttp_retry import RetryClient, ExponentialRetry
from fastapi import FastAPI, BackgroundTasks
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import column_index_from_string
from PIL import Image as PILImage
from PIL import UnidentifiedImageError, ImageOps
from tldextract import tldextract
import urllib.parse
from pathlib import Path
import numpy as np
from collections import Counter
import re
import random
from functools import partial
from app_config import engine, conn_str
from email_utils import send_email, send_message_email
from s3_utils import upload_file_to_space
from openpyxl.utils.units import pixels_to_EMU, points_to_pixels
# --- Global Setup ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
app = FastAPI()

# --- Constants ---
MAX_THREADS = int(os.environ.get('MAX_THREADS', 10))
VALID_IMAGE_TYPES = ['image/jpeg', 'image/png', 'image/gif', 'image/bmp', 'image/webp', 'image/avif', 'image/tiff', 'image/x-icon']
MIN_IMAGE_SIZE = 1000  # Minimum content length in bytes
MAX_IMAGE_DIMENSION = 130  # For resizing
MIN_DIMENSION_FOR_BORDER = 10  # Minimum dimension to attempt border removal
BORDER_CROP_WIDTH = 5  # Pixels to check for border on each side
BORDER_UNIFORMITY_THRESHOLD = 0.05  # 5% of pixels must match dominant color
BORDER_COLOR_TOLERANCE = 10  # RGB tolerance for border color matching
EMU_PER_PIXEL = 9525  # EMUs per pixel at 96 DPI
def points_to_pixels(points):
    """Convert points to pixels assuming 96 DPI (1 point = 1.333 pixels)."""
    return points * 1.333
def pixels_to_EMU(pixels: float) -> int:
    """Convert pixels to English Metric Units (EMU) using 96 DPI."""
    return int(round(pixels * EMU_PER_PIXEL))
def get_user_agents(logger_instance: logging.Logger) -> List[str]:
    """
    Fetches a list of user agents from the DataProxy API.
    Returns a default list if the API call fails.
    """
    default_ua = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"]
    try:
        api_url = "https://api.thedataproxy.com/v2/user-agents/?skip=0&limit=100"
        logger_instance.info(f"Fetching user agents from {api_url}")
        response = requests.get(api_url, timeout=15)
        response.raise_for_status()
        data = response.json()
        user_agents = [ua.get("user_agent") for ua in data.get("data", []) if ua.get("user_agent")]
        if not user_agents:
            logger_instance.warning("User-Agent API returned no agents. Using default.")
            return default_ua
        logger_instance.info(f"Successfully fetched {len(user_agents)} user agents.")
        return user_agents
    except requests.exceptions.RequestException as e:
        logger_instance.error(f"Failed to fetch user agents from API: {e}. Using default.")
        return default_ua

# --- General Utility Functions ---
def setup_logging(file_id: str, timestamp: str) -> logging.Logger:
    """Configure logging to save logs to a file for a specific job run."""
    log_dir = Path('jobs') / file_id / timestamp / 'logs'
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / f'job_{file_id}_{timestamp}.log'
    
    logger_instance = logging.getLogger(f"job_{file_id}_{timestamp}")
    logger_instance.setLevel(logging.INFO)
    
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    
    if not logger_instance.handlers:
        file_handler = logging.FileHandler(log_file)
        file_handler.setFormatter(formatter)
        
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(formatter)
        
        logger_instance.addHandler(file_handler)
        logger_instance.addHandler(console_handler)
        logger_instance.propagate = False
    
    logger_instance.info(f"Logging initialized for FileID: {file_id}, Timestamp: {timestamp}")
    return logger_instance

async def create_temp_dirs(unique_id: str) -> Tuple[str, str]:
    """Create temporary directories for a job."""
    base_dir = Path.cwd() / 'temp_files'
    temp_images_dir = base_dir / 'images' / unique_id
    temp_excel_dir = base_dir / 'excel' / unique_id
    
    loop = asyncio.get_running_loop()
    await loop.run_in_executor(None, partial(os.makedirs, temp_images_dir, exist_ok=True))
    await loop.run_in_executor(None, partial(os.makedirs, temp_excel_dir, exist_ok=True))
    
    return str(temp_images_dir), str(temp_excel_dir)

async def cleanup_temp_dirs(directories: List[str]):
    """Remove temporary directories."""
    loop = asyncio.get_running_loop()
    for dir_path in directories:
        await loop.run_in_executor(None, partial(shutil.rmtree, dir_path, ignore_errors=True))

# --- Database Functions ---
def get_file_type_id(file_id: int, logger_instance: logging.Logger) -> Optional[int]:
    """Retrieve the FileTypeID from the database."""
    with pyodbc.connect(conn_str) as connection:
        cursor = connection.cursor()
        query = "SELECT FileTypeID FROM utb_ImageScraperFiles WHERE ID = ?"
        cursor.execute(query, (file_id,))
        result = cursor.fetchone()
        logger_instance.info(f"FileTypeID for {file_id} is {result[0] if result else 'Not Found'}")
        return result[0] if result else None

def get_file_location_and_header(file_id: int, logger_instance: logging.Logger) -> Tuple[str, Optional[int], Optional[str]]:
    """Retrieve the source file location URL, header row index, and user email from the database."""
    with pyodbc.connect(conn_str) as connection:
        cursor = connection.cursor()
        query = "SELECT FileLocationUrl, UserHeaderIndex, UserEmail FROM utb_ImageScraperFiles WHERE ID = ?"
        cursor.execute(query, (file_id,))
        result = cursor.fetchone()
        if not result:
            raise FileNotFoundError(f"No file location found in DB for FileID {file_id}")
        file_location = result[0]
        header_row = result[1]
        user_email = result[2]
        if header_row is not None:
            try:
                header_row = int(header_row)
            except (ValueError, TypeError):
                logger_instance.warning(f"Invalid UserHeaderIndex for FileID {file_id}: {header_row}. Setting to None.")
                header_row = None
        logger_instance.info(f"FileLocationUrl for FileID {file_id}: {file_location}")
        logger_instance.info(f"UserHeaderIndex for FileID {file_id}: {header_row if header_row is not None else 'Not Found'}")
        logger_instance.info(f"UserEmail for FileID {file_id}: {user_email}")
        return file_location, header_row, user_email

def get_images_excel_db(file_id: int, logger_instance: logging.Logger) -> pd.DataFrame:
    """
    Retrieve detailed image and product data for Excel processing.
    Includes all rows, even those without image results, using LEFT JOIN.
    """
    with pyodbc.connect(conn_str) as connection:
        cursor = connection.cursor()
        cursor.execute("UPDATE utb_ImageScraperFiles SET CreateFileStartTime = GETDATE() WHERE ID = ?", (file_id,))
        connection.commit()
    
    query = """
        SELECT
            s.ExcelRowID, r.ImageUrl, r.ImageUrlThumbnail, r.SortOrder,
            s.ProductBrand AS Brand, s.ProductModel AS Style,
            s.ProductColor AS Color, s.ProductCategory AS Category,
            s.ProductMSRP AS MSRP
        FROM utb_ImageScraperFiles f
        INNER JOIN utb_ImageScraperRecords s ON s.FileID = f.ID 
        LEFT JOIN utb_ImageScraperResult r ON r.EntryID = s.EntryID AND r.SortOrder > 0
        WHERE f.ID = ?
        ORDER BY s.ExcelRowID, r.SortOrder
    """
    
    logger_instance.info(f"Executing query to fetch all records for FileID: {file_id}")
    return pd.read_sql_query(query, engine, params=[(file_id,)])

def update_file_location_complete(file_id: int, file_location: str, logger_instance: logging.Logger):
    """Update the completed file location URL in the database."""
    with pyodbc.connect(conn_str) as connection:
        cursor = connection.cursor()
        query = "UPDATE utb_ImageScraperFiles SET FileLocationURLComplete = ? WHERE ID = ?"
        cursor.execute(query, (file_location, file_id))
        connection.commit()
        logger_instance.info(f"Updated completed file location for FileID {file_id}")

# --- Image Downloading Functions ---
async def validate_image_response(response, url: str, logger_instance: logging.Logger) -> Tuple[bool, Optional[bytes]]:
    if response.status != 200:
        logger_instance.warning(f"Invalid status code {response.status} for URL: {url}")
        return False, None
    if not any(response.headers.get('Content-Type', '').lower().startswith(it) for it in VALID_IMAGE_TYPES):
        logger_instance.warning(f"Invalid Content-Type for URL: {url}")
        return False, None
    if int(response.headers.get('Content-Length', 0)) < MIN_IMAGE_SIZE:
        logger_instance.warning(f"Content too small for URL: {url}")
        return False, None
    return True, await response.read()

async def image_download(semaphore, item: Dict, save_path: str, session, logger_instance: logging.Logger, user_agents: List[str]):
    async with semaphore:
        row_id = item['ExcelRowID']
        image_name = str(row_id)
        loop = asyncio.get_running_loop()

        for i, (url, thumb_url, sort_order) in enumerate(item['image_options']):

            async def attempt_download(download_url: str, is_thumb: bool) -> bool:
                if not download_url: return False
                try:
                    headers = {"User-Agent": random.choice(user_agents)}
                    async with session.get(download_url, headers=headers) as response:
                        is_valid, data = await validate_image_response(response, download_url, logger_instance)
                        if not is_valid: return False

                        final_path = os.path.join(save_path, f"{image_name}.png")

                        def save_and_verify_image_sync():
                            try:
                                # Save the image
                                with PILImage.open(BytesIO(data)) as img:
                                    # Ensure proper mode conversion during save
                                    if img.mode == 'RGBA':
                                        background = PILImage.new('RGB', img.size, (255, 255, 255))
                                        background.paste(img, mask=img.split()[3])
                                        img = background
                                    elif img.mode != 'RGB':
                                        img = img.convert('RGB')
                                    img.save(final_path, 'PNG', compress_level=6)

                                # Immediately verify the saved image
                                if not os.path.exists(final_path):
                                    logger_instance.warning(f"Image file not created: {final_path}")
                                    return False

                                file_size = os.path.getsize(final_path)
                                if file_size < MIN_IMAGE_SIZE:
                                    logger_instance.warning(f"Saved image too small ({file_size} bytes): {final_path}")
                                    try:
                                        os.remove(final_path)
                                    except:
                                        pass
                                    return False

                                # Try to open and verify the image
                                with PILImage.open(final_path) as verify_img:
                                    verify_img.verify()

                                # Load it again to ensure it's readable
                                with PILImage.open(final_path) as verify_img2:
                                    verify_img2.load()
                                    width, height = verify_img2.size
                                    if width <= 0 or height <= 0:
                                        logger_instance.warning(f"Invalid image dimensions {width}x{height}: {final_path}")
                                        try:
                                            os.remove(final_path)
                                        except:
                                            pass
                                        return False

                                return True
                            except Exception as verify_error:
                                logger_instance.warning(f"Image verification failed: {verify_error}")
                                # Clean up corrupt file
                                if os.path.exists(final_path):
                                    try:
                                        os.remove(final_path)
                                    except:
                                        pass
                                return False

                        is_verified = await loop.run_in_executor(None, save_and_verify_image_sync)

                        if is_verified:
                            logger_instance.info(f"SUCCESS (SortOrder {sort_order}) for Row {row_id} from {'thumbnail' if is_thumb else 'main'} URL.")
                            return True
                        else:
                            logger_instance.warning(f"Image corrupt/invalid for Row {row_id} from {'thumbnail' if is_thumb else 'main'} URL, trying next option...")
                            return False

                except Exception as e:
                    logger_instance.warning(f"Download attempt {i+1} failed for Row {row_id} ({'thumb' if is_thumb else 'main'}): {e}")
                    return False

            if await attempt_download(url, is_thumb=False):
                return True
            if await attempt_download(thumb_url, is_thumb=True):
                return True

        logger_instance.error(f"All download attempts FAILED for Row {row_id}. No more images to try.")
        return False

async def download_all_images(data: List[Dict], save_path: str, logger_instance: logging.Logger, user_agents: List[str]):
    """Download all images concurrently, trying fallbacks for each item."""
    domains = []
    for item in data:
        if item.get('image_options'):
            first_url = item['image_options'][0][0]
            if first_url:
                domains.append(tldextract.extract(first_url).top_domain_under_public_suffix)
    
    pool_size = min(500, max(10, len(Counter(domains)) * 2))
    timeout_config = ClientTimeout(total=60)
    retry_options = ExponentialRetry(attempts=3, start_timeout=3)
    
    async with RetryClient(raise_for_status=False, retry_options=retry_options, timeout=timeout_config) as session:
        semaphore = asyncio.Semaphore(pool_size)
        tasks = [image_download(semaphore, item, save_path, session, logger_instance, user_agents) for item in data]
        results = await asyncio.gather(*tasks)
        
        failed_count = len([res for res in results if not res])
        if failed_count > 0:
            logger_instance.warning(f"{failed_count} product rows failed to download any image.")

# --- Image & Excel Processing Functions ---
def resize_image(image_path: str, logger_instance: logging.Logger) -> bool:
    try:
        # Check if file exists first
        if not os.path.exists(image_path):
            logger_instance.warning(f"Image file does not exist: {image_path}")
            return False

        with PILImage.open(image_path) as img:
            img.load()  # Ensure image is fully loaded

            # Resize in memory (upscale or downscale)
            width, height = img.size
            ratio = min(MAX_IMAGE_DIMENSION / width, MAX_IMAGE_DIMENSION / height)
            new_w = int(width * ratio)
            new_h = int(height * ratio)
            
            if (new_w, new_h) != (width, height):
                img = img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)
                logger_instance.info(f"Resized image {image_path} from {width}x{height} to {new_w}x{new_h} (max dim {MAX_IMAGE_DIMENSION})")

            # Save once at the end
            img.save(image_path, 'PNG', compress_level=6, quality=95)
            return True
    except Exception as e:
        logger_instance.error(f"Error resizing image {image_path}: {e}", exc_info=True)
        return False

def get_last_non_empty_row(ws, column: str, header_row: int, logger_instance: logging.Logger) -> int:
    """Find the last non-empty row in the specified column, starting after header_row."""
    last_row = header_row
    for row in ws[f"{column}{header_row + 1}:{column}{ws.max_row}"]:
        if row[0].value is not None and str(row[0].value).strip():
            last_row = max(last_row, row[0].row)
    logger_instance.info(f"Last non-empty row in column {column}: {last_row}")
    return last_row

def verify_and_process_image(image_path: str, logger_instance: logging.Logger) -> bool:
    """Verify, crop, and process image in memory to avoid corruption."""
    img = None
    backup_path = None

    try:
        # Check if file exists first
        if not os.path.exists(image_path):
            logger_instance.warning(f"Image file does not exist: {image_path}")
            return False

        # Verify image integrity
        with PILImage.open(image_path) as temp_img:
            temp_img.verify()

        original_size = os.path.getsize(image_path)
        if original_size < MIN_IMAGE_SIZE:
            logger_instance.warning(f"File too small: {image_path}")
            return False

        # Create backup of original image before processing
        backup_path = f"{image_path}.backup"
        shutil.copy2(image_path, backup_path)

        # Load image for processing
        img = PILImage.open(image_path)
        img.load()

        # Convert to RGB if necessary to ensure compatibility
        if img.mode not in ('RGB', 'RGBA'):
            logger_instance.info(f"Converting image mode from {img.mode} to RGB: {image_path}")
            img = img.convert('RGB')

        # Process image (cropping and border removal) with error handling
        try:
            img = process_image_remove_lines(image_path, img, logger_instance)
        except Exception as process_error:
            logger_instance.warning(f"Image processing failed, using simple resize instead: {process_error}")
            # Fallback: just reload the original image
            img.close()
            img = PILImage.open(backup_path)
            img.load()
            if img.mode == 'RGBA':
                background = PILImage.new('RGB', img.size, (255, 255, 255))
                background.paste(img, mask=img.split()[3])
                img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')

        # Resize to fit max dimension (upscale or downscale)
        w, h = img.size

        # Ensure dimensions are valid
        if w <= 0 or h <= 0:
            logger_instance.error(f"Invalid image dimensions: {w}x{h} for {image_path}")
            return False

        ratio = min(MAX_IMAGE_DIMENSION / w, MAX_IMAGE_DIMENSION / h)
        new_w = max(1, int(w * ratio))
        new_h = max(1, int(h * ratio))

        if (new_w, new_h) != (w, h):
            img = img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)
            logger_instance.info(f"Resized image {image_path} from {w}x{h} to {new_w}x{new_h} (max dim {MAX_IMAGE_DIMENSION})")

        # Ensure RGB mode before saving
        if img.mode != 'RGB':
            img = img.convert('RGB')

        # Save processed image with error handling
        try:
            img.save(image_path, 'PNG', compress_level=6, optimize=True)
            logger_instance.info(f"Saved image with compression_level=6, optimize=True")
        except Exception as save_error:
            logger_instance.error(f"Failed to save processed image: {save_error}. Trying with lower compression.")
            # Fallback: save with minimal compression
            try:
                img.save(image_path, 'PNG', compress_level=1)
                logger_instance.info(f"Saved image with compression_level=1 (fallback)")
            except Exception as fallback_save_error:
                logger_instance.error(f"Fallback save also failed: {fallback_save_error}")
                return False

        # Verify the saved image is valid
        if not os.path.exists(image_path):
            logger_instance.error(f"Image file disappeared after saving: {image_path}")
            # Restore from backup
            if backup_path and os.path.exists(backup_path):
                shutil.copy2(backup_path, image_path)
                logger_instance.info(f"Restored image from backup: {image_path}")
            return False

        file_size = os.path.getsize(image_path)
        if file_size < MIN_IMAGE_SIZE:
            logger_instance.error(f"Saved image file too small ({file_size} bytes): {image_path}")
            # Restore from backup
            if backup_path and os.path.exists(backup_path):
                shutil.copy2(backup_path, image_path)
                logger_instance.info(f"Restored image from backup due to small size: {image_path}")
            return False

        # Try to re-open the saved image to ensure it's valid
        try:
            with PILImage.open(image_path) as verify_img:
                verify_img.verify()
            # Load it again to ensure it can be fully read
            with PILImage.open(image_path) as verify_img2:
                verify_img2.load()
        except Exception as verify_error:
            logger_instance.error(f"Saved image failed verification: {image_path}, error: {verify_error}")
            # Restore from backup and try simple processing
            if backup_path and os.path.exists(backup_path):
                logger_instance.info(f"Attempting simple processing fallback for: {image_path}")
                try:
                    with PILImage.open(backup_path) as fallback_img:
                        fallback_img.load()
                        if fallback_img.mode == 'RGBA':
                            background = PILImage.new('RGB', fallback_img.size, (255, 255, 255))
                            background.paste(fallback_img, mask=fallback_img.split()[3])
                            fallback_img = background
                        elif fallback_img.mode != 'RGB':
                            fallback_img = fallback_img.convert('RGB')

                        # Just resize, no other processing
                        w, h = fallback_img.size
                        ratio = min(MAX_IMAGE_DIMENSION / w, MAX_IMAGE_DIMENSION / h)
                        new_w = max(1, int(w * ratio))
                        new_h = max(1, int(h * ratio))
                        if (new_w, new_h) != (w, h):
                            fallback_img = fallback_img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)

                        fallback_img.save(image_path, 'PNG', compress_level=1)
                        logger_instance.info(f"Successfully saved with fallback processing: {image_path}")
                except Exception as fallback_error:
                    logger_instance.error(f"Fallback processing also failed: {fallback_error}")
                    return False
            else:
                return False

        # Clean up backup if successful
        if backup_path and os.path.exists(backup_path):
            try:
                os.remove(backup_path)
            except:
                pass

        logger_instance.info(f"Image verified and processed: {image_path}")
        return True

    except Exception as e:
        logger_instance.error(f"Image verification failed for {image_path}: {e}", exc_info=True)
        # Try to restore from backup
        if backup_path and os.path.exists(backup_path):
            try:
                shutil.copy2(backup_path, image_path)
                logger_instance.info(f"Restored original image from backup after error: {image_path}")
            except:
                pass
        return False
    finally:
        # Ensure image is closed to prevent file handle leaks
        if img is not None:
            try:
                img.close()
            except:
                pass
        # Clean up backup file
        if backup_path and os.path.exists(backup_path):
            try:
                os.remove(backup_path)
            except:
                pass

def process_image_remove_lines(image_path: str, img: PILImage.Image, logger_instance: logging.Logger) -> PILImage.Image:
    """Process image in memory to remove uniform lines and borders, preventing smearing."""
    try:
        img = ImageOps.exif_transpose(img.copy())
        arr = np.array(img)
        original_height, original_width, _ = arr.shape

        # Detect non-uniform horizontal rows
        valid_row_indices = [y for y in range(original_height) if np.any(np.max(arr[y], axis=0) - np.min(arr[y], axis=0) >= 15)]
        row_cropped = False
        if valid_row_indices:
            rows_to_keep = get_valid_indices_with_padding(valid_row_indices, original_height - 1, pad=5)
            row_cropped = len(rows_to_keep) != original_height
            arr = arr[rows_to_keep]
            logger_instance.info(f"Kept {len(rows_to_keep)} of {original_height} rows for {image_path}")
        else:
            logger_instance.info(f"No non-uniform rows found in {image_path}")

        # Detect non-uniform vertical columns
        arr_t = np.transpose(arr, (1, 0, 2))
        new_width = arr_t.shape[0]
        valid_col_indices = [x for x in range(new_width) if np.any(np.max(arr_t[x], axis=0) - np.min(arr_t[x], axis=0) >= 15)]
        col_cropped = False
        if valid_col_indices:
            cols_to_keep = get_valid_indices_with_padding(valid_col_indices, new_width - 1, pad=5)
            col_cropped = len(cols_to_keep) != new_width
            arr_t = arr_t[cols_to_keep]
            arr = np.transpose(arr_t, (1, 0, 2))
            logger_instance.info(f"Kept {len(cols_to_keep)} of {new_width} columns for {image_path}")
        else:
            logger_instance.info(f"No non-uniform columns found in {image_path}")

        # Check if cropped image is too small
        cleaned_height, cleaned_width, _ = arr.shape
        min_dim = 20
        if (row_cropped or col_cropped) and (cleaned_height < min_dim or cleaned_width < min_dim):
            logger_instance.warning(f"Cropped image too small ({cleaned_height}x{cleaned_width}) for {image_path}. Using original.")
            return img

        # Border removal (only if dimensions are sufficient)
        pixels = arr
        if cleaned_height >= MIN_DIMENSION_FOR_BORDER and cleaned_width >= MIN_DIMENSION_FOR_BORDER:
            try:
                # Collect border pixels from outer 5 pixels, avoiding corner overlaps
                top_border = pixels[:BORDER_CROP_WIDTH, BORDER_CROP_WIDTH:-BORDER_CROP_WIDTH, :]
                bottom_border = pixels[-BORDER_CROP_WIDTH:, BORDER_CROP_WIDTH:-BORDER_CROP_WIDTH, :]
                left_border = pixels[BORDER_CROP_WIDTH:-BORDER_CROP_WIDTH, :BORDER_CROP_WIDTH, :]
                right_border = pixels[BORDER_CROP_WIDTH:-BORDER_CROP_WIDTH, -BORDER_CROP_WIDTH:, :]
                
                border_pixels = np.vstack([
                    top_border.reshape(-1, 3),
                    bottom_border.reshape(-1, 3),
                    left_border.reshape(-1, 3),
                    right_border.reshape(-1, 3)
                ])
                
                # Get dominant border color
                colors, counts = np.unique(border_pixels, axis=0, return_counts=True)
                total_border_pixels = len(border_pixels)
                dominant_color = colors[np.argmax(counts)]
                dominant_fraction = counts.max() / total_border_pixels
                
                # Check if border is light and sufficiently uniform
                if np.sum(dominant_color) > 750 and dominant_fraction > BORDER_UNIFORMITY_THRESHOLD:
                    # Create mask for border areas only
                    mask = np.zeros((cleaned_height, cleaned_width), dtype=bool)
                    mask[:BORDER_CROP_WIDTH, :] = np.all(np.abs(pixels[:BORDER_CROP_WIDTH, :, :] - dominant_color) <= BORDER_COLOR_TOLERANCE, axis=2)
                    mask[-BORDER_CROP_WIDTH:, :] = np.all(np.abs(pixels[-BORDER_CROP_WIDTH:, :, :] - dominant_color) <= BORDER_COLOR_TOLERANCE, axis=2)
                    mask[:, :BORDER_CROP_WIDTH] = np.all(np.abs(pixels[:, :BORDER_CROP_WIDTH, :] - dominant_color) <= BORDER_COLOR_TOLERANCE, axis=2)
                    mask[:, -BORDER_CROP_WIDTH:] = np.all(np.abs(pixels[:, -BORDER_CROP_WIDTH:, :] - dominant_color) <= BORDER_COLOR_TOLERANCE, axis=2)
                    
                    # Avoid modifying inner region
                    mask[BORDER_CROP_WIDTH:-BORDER_CROP_WIDTH, BORDER_CROP_WIDTH:-BORDER_CROP_WIDTH] = False
                    
                    # Replace border pixels with white
                    pixels[mask] = np.array([255, 255, 255])
                    arr = pixels
                    logger_instance.info(f"Applied border removal for {image_path} (dominant color {dominant_color}, fraction {dominant_fraction:.2f})")
                else:
                    logger_instance.info(f"No light uniform border detected for {image_path} (fraction {dominant_fraction:.2f}, color sum {np.sum(dominant_color)})")
            except Exception as e:
                logger_instance.warning(f"Border removal failed for {image_path}: {e}. Using cropped image.")
        else:
            logger_instance.info(f"Image too small for border removal ({cleaned_height}x{cleaned_width}) for {image_path}")

        return PILImage.fromarray(arr.astype(np.uint8))
    except Exception as e:
        logger_instance.error(f"Error processing image {image_path} to remove lines: {e}", exc_info=True)
        return img

def get_valid_indices_with_padding(valid_indices, max_index, pad=5):
    """Identify indices with padding for non-uniform rows/columns."""
    padded = set()
    for idx in valid_indices:
        start = max(0, idx - pad)
        end = min(max_index, idx + pad + 1)
        padded.update(range(start, end))
    return sorted(padded)

def fix_excel_image_paths(excel_file: str, logger_instance: logging.Logger) -> bool:
    """
    Fix openpyxl 3.1.5 bug where image relationship paths are absolute instead of relative.
    Post-processes the saved Excel file to change /xl/media/ to ../media/ in drawing relationships.
    """
    try:
        logger_instance.info(f"Fixing image paths in: {excel_file}")

        # Create a temporary file for the fixed version
        temp_file = excel_file + ".tmp"

        # Open the Excel file as a ZIP
        with zipfile.ZipFile(excel_file, 'r') as zip_read:
            with zipfile.ZipFile(temp_file, 'w', zipfile.ZIP_DEFLATED) as zip_write:
                for item in zip_read.infolist():
                    data = zip_read.read(item.filename)

                    # Fix the drawing relationships file
                    if item.filename.endswith('drawings/_rels/drawing1.xml.rels'):
                        data_str = data.decode('utf-8')
                        # Replace absolute paths with relative paths
                        data_str = data_str.replace('Target="/xl/media/', 'Target="../media/')
                        data = data_str.encode('utf-8')
                        logger_instance.info(f"Fixed image paths in {item.filename}")

                    zip_write.writestr(item, data)

        # Replace the original file with the fixed one
        shutil.move(temp_file, excel_file)
        logger_instance.info(f"Successfully fixed image paths in: {excel_file}")
        return True

    except Exception as e:
        logger_instance.error(f"Failed to fix image paths in {excel_file}: {e}", exc_info=True)
        # Clean up temp file if it exists
        if os.path.exists(temp_file):
            os.remove(temp_file)
        return False


def write_excel_distro(local_filename: str, temp_dir: str, image_data: List[Dict], header_row: int, logger_instance: logging.Logger, row_offset: int = 0):
    try:
        logger_instance.info(f"=== Starting write_excel_distro ===")
        logger_instance.info(f"Template file: {local_filename}, size: {os.path.getsize(local_filename)} bytes")

        # Clean the template file (mark external links as broken to preserve formulas)
        logger_instance.info("Cleaning template file (marking external links as broken)...")
        original_filename = local_filename
        cleaned_filename = clean_template_file(local_filename, logger_instance, mark_broken=True)
        logger_instance.info(f"Using cleaned template: {cleaned_filename}")

        # Validate template file BEFORE loading
        logger_instance.info("Validating template file before processing...")
        if not validate_excel_file(cleaned_filename, logger_instance):
            logger_instance.warning("Template file has validation warnings, but proceeding...")

        header_row = int(header_row)
        if header_row < 0:
            logger_instance.error(f"Invalid header_row {header_row}. Setting to 0.")
            header_row = 0
        if row_offset < 0:
            logger_instance.warning(f"Negative row_offset {row_offset} provided. Ensure this is intentional.")

        logger_instance.info(f"Loading workbook from: {cleaned_filename}")

        # Load the workbook
        wb = load_workbook(cleaned_filename)
        logger_instance.info(f"Workbook loaded successfully")

        ws = wb.active
        logger_instance.info(f"Active worksheet: {ws.title}, dimensions: {ws.max_row}x{ws.max_column}")

        # CRITICAL FIX: Clear ALL existing images to prevent TwoCellAnchor/OneCellAnchor conflicts
        if hasattr(ws, '_images') and ws._images:
            existing_count = len(ws._images)
            ws._images = []
            logger_instance.info(f"Cleared {existing_count} existing images from template to prevent anchor conflicts")

        image_map = {}
        for path_obj in Path(temp_dir).iterdir():
            if not path_obj.is_file():
                continue
            stem = path_obj.stem
            if stem.isdigit():
                image_map[int(stem)] = str(path_obj)

        row_dim = ws.row_dimensions.get(header_row + 1)
        if row_dim is not None and getattr(row_dim, "height", None) is not None:
            DEFAULT_ROW_HEIGHT_POINTS = row_dim.height
        else:
            DEFAULT_ROW_HEIGHT_POINTS = 12.75
        logger_instance.info(f"Using template row height: {DEFAULT_ROW_HEIGHT_POINTS} points from row {header_row + 1}")

        row_data_map = {item['ExcelRowID']: item for item in image_data}
        last_non_empty_row = get_last_non_empty_row(ws, column='B', header_row=header_row, logger_instance=logger_instance)

        data_row_ids = sorted(row_data_map.keys())
        if data_row_ids:
            min_row_id = data_row_ids[0]
            max_row_id = data_row_ids[-1]
        else:
            min_row_id = 1
            relative_last_row = last_non_empty_row - (header_row + 1)
            relative_last_row = max(1, relative_last_row)
            max_row_id = relative_last_row
            logger_instance.warning(
                f"No data in image_data, deriving row range 1-{max_row_id} from template last row {last_non_empty_row}"
            )

        base_row = header_row + row_offset + 2
        if base_row < 1:
            logger_instance.error(f"Computed base_row {base_row} is less than 1. Aborting.")
            raise ValueError("Invalid row range: base_row is less than 1")

        # CRITICAL FIX: Only create rows that have actual data, not the entire range
        # Create mapping of ExcelRowID -> Excel row number
        row_id_to_row_num = {}
        for idx, row_id in enumerate(data_row_ids):
            row_id_to_row_num[row_id] = base_row + idx

        # Calculate max row needed (only for rows with data)
        max_needed_row = base_row + len(data_row_ids) - 1 if data_row_ids else base_row
        max_needed_row = max(max_needed_row, last_non_empty_row)

        logger_instance.info(
            f"Row mapping summary -> base_row: {base_row}, min_row_id: {min_row_id}, max_row_id: {max_row_id}, "
            f"data_rows_count: {len(data_row_ids)}, max_needed_row: {max_needed_row}"
        )

        if ws.max_row < max_needed_row:
            logger_instance.info(f"Appending {max_needed_row - ws.max_row} rows to worksheet")
            for row_num in range(ws.max_row + 1, max_needed_row + 1):
                ws.append([''] * ws.max_column)
                ws.row_dimensions[row_num].height = DEFAULT_ROW_HEIGHT_POINTS

        CELL_WIDTH_POINTS = 100
        CELL_HEIGHT_POINTS = max(DEFAULT_ROW_HEIGHT_POINTS, 150)
        PADDING_POINTS = 5

        # Process only rows that actually have data
        for row_id in data_row_ids:
            row_num = row_id_to_row_num[row_id]

            ws.row_dimensions[row_num].height = CELL_HEIGHT_POINTS

            image_path = image_map.get(row_id)
            if image_path:
                logger_instance.info(f"DEBUG: Processing image for Row {row_id}: {image_path}")
                if not verify_and_process_image(image_path, logger_instance):
                    error_msg = f"DEBUG: verify_and_process_image FAILED for Row {row_id} at {image_path}"
                    logger_instance.error(error_msg)
                    raise ValueError(error_msg)

                try:
                    # Validate image file one more time before inserting
                    if not os.path.exists(image_path):
                        error_msg = f"DEBUG: Image file DOES NOT EXIST after verify_and_process for Row {row_id}: {image_path}"
                        logger_instance.error(error_msg)
                        raise FileNotFoundError(error_msg)

                    file_size = os.path.getsize(image_path)
                    if file_size < MIN_IMAGE_SIZE:
                        error_msg = f"DEBUG: Image file TOO SMALL after verify_and_process for Row {row_id}: {file_size} bytes at {image_path}"
                        logger_instance.error(error_msg)
                        raise ValueError(error_msg)

                    # CRITICAL: Validate PIL can read the image and get dimensions BEFORE creating openpyxl Image
                    try:
                        pil_img = PILImage.open(image_path)
                        pil_width, pil_height = pil_img.size
                        pil_img.close()
                        if pil_width <= 0 or pil_height <= 0:
                            error_msg = f"DEBUG: PIL image has invalid dimensions for Row {row_id}: {pil_width}x{pil_height}"
                            logger_instance.error(error_msg)
                            raise ValueError(error_msg)
                        logger_instance.info(f"DEBUG: PIL validated image dimensions for Row {row_id}: {pil_width}x{pil_height}")
                    except Exception as pil_error:
                        error_msg = f"DEBUG: PIL failed to read image for Row {row_id}: {pil_error}"
                        logger_instance.error(error_msg)
                        raise ValueError(error_msg) from pil_error

                    # Try to create openpyxl Image object with error handling
                    logger_instance.info(f"DEBUG: Creating openpyxl Image object for Row {row_id}")
                    img = Image(image_path)
                    logger_instance.info(f"Image object created successfully for Row {row_id}")
                    img_height_pixels = getattr(img, 'height', 0)
                    img_width_pixels = getattr(img, 'width', 0)

                    # Validate dimensions
                    if img_height_pixels <= 0 or img_width_pixels <= 0:
                        error_msg = f"DEBUG: Invalid image dimensions for Row {row_id}: {img_width_pixels}x{img_height_pixels}"
                        logger_instance.error(error_msg)
                        raise ValueError(error_msg)

                    img_height_points = img_height_pixels * 72 / 96
                    img_width_points = img_width_pixels / 7

                    required_width = img_width_points + PADDING_POINTS * 2
                    current_width = ws.column_dimensions['A'].width or 8.43
                    ws.column_dimensions['A'].width = min(CELL_WIDTH_POINTS, max(current_width, required_width))

                    cell_width_pixels = ws.column_dimensions['A'].width * 7
                    cell_height_pixels = points_to_pixels(CELL_HEIGHT_POINTS)
                    x_offset_pixels = max(0, (cell_width_pixels - img_width_pixels) / 2)
                    y_offset_pixels = max(0, (cell_height_pixels - img_height_pixels) / 2)

                    width_emu = pixels_to_EMU(img_width_pixels)
                    height_emu = pixels_to_EMU(img_height_pixels)

                    # Validate EMU values to prevent XML corruption
                    if width_emu <= 0 or height_emu <= 0:
                        error_msg = f"DEBUG: Invalid EMU dimensions for Row {row_id}: {width_emu}x{height_emu}"
                        logger_instance.error(error_msg)
                        raise ValueError(error_msg)
                    elif width_emu > 10000000 or height_emu > 10000000:  # Max ~1050 pixels
                        error_msg = f"DEBUG: EMU dimensions too large for Row {row_id}: {width_emu}x{height_emu}"
                        logger_instance.error(error_msg)
                        raise ValueError(error_msg)

                    anchor_row = max(0, row_num - 1)
                    x_offset_emu = pixels_to_EMU(x_offset_pixels)
                    y_offset_emu = pixels_to_EMU(y_offset_pixels)

                    # Ensure offsets are non-negative and within valid range
                    # Max offset is approximately 9,525,000 EMU (1000 pixels)
                    MAX_OFFSET_EMU = 9525000
                    x_offset_emu = max(0, min(x_offset_emu, MAX_OFFSET_EMU))
                    y_offset_emu = max(0, min(y_offset_emu, MAX_OFFSET_EMU))

                    logger_instance.info(
                        f"Creating anchor for Row {row_id} at Excel row {row_num}: "
                        f"marker(col=0, colOff={x_offset_emu}, row={anchor_row}, rowOff={y_offset_emu}), "
                        f"size({width_emu}x{height_emu} EMU = {img_width_pixels}x{img_height_pixels} px)"
                    )
                    marker = AnchorMarker(
                        col=0,
                        colOff=x_offset_emu,
                        row=anchor_row,
                        rowOff=y_offset_emu
                    )
                    img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(width_emu, height_emu))

                    # CRITICAL: Final validation before adding to worksheet
                    # Ensure Image object still has valid dimensions after anchor is set
                    final_width = getattr(img, 'width', 0)
                    final_height = getattr(img, 'height', 0)
                    if final_width <= 0 or final_height <= 0:
                        error_msg = f"CRITICAL: Image dimensions became 0 after anchor was set for Row {row_id}: {final_width}x{final_height}"
                        logger_instance.error(error_msg)
                        raise ValueError(error_msg)

                    image_count_before = len(ws._images) if hasattr(ws, '_images') else 0
                    logger_instance.info(f"Adding image to worksheet (current count: {image_count_before})")
                    ws.add_image(img)
                    image_count_after = len(ws._images) if hasattr(ws, '_images') else 0
                    logger_instance.info(f"Image added successfully (new count: {image_count_after})")

                    logger_instance.info(
                        f"✓ Added image for Row {row_id} at Excel row {row_num}, "
                        f"height={CELL_HEIGHT_POINTS} points, width={ws.column_dimensions['A'].width:.2f} points"
                    )
                except Exception as img_error:
                    logger_instance.error(f"DEBUG: EXCEPTION while inserting image for Row {row_id}: {img_error}", exc_info=True)
                    raise  # Re-raise to stop execution for debugging
            else:
                logger_instance.info(f"No image found for Row {row_id}, writing metadata only")

            item = row_data_map.get(row_id)
            if item:
                if row_num > header_row + 1:
                    ws[f"B{row_num}"] = item.get('Brand', '')
                    ws[f"D{row_num}"] = item.get('Style', '')
                    ws[f"E{row_num}"] = item.get('Color', '')
                    ws[f"H{row_num}"] = item.get('Category', '')
                    logger_instance.info(f"Wrote metadata for Row {row_id} at Excel row {row_num}")
                else:
                    logger_instance.warning(f"Skipping metadata write for row {row_num} as it is the header row")
            else:
                if row_num > header_row + 1:
                    ws[f"B{row_num}"] = ''
                    ws[f"D{row_num}"] = ''
                    ws[f"E{row_num}"] = ''
                    ws[f"H{row_num}"] = ''
                    logger_instance.info(f"Filled missing row {row_num} (ExcelRowID {row_id}) with empty metadata")
                else:
                    logger_instance.info(f"Skipping row {row_num} (ExcelRowID {row_id}) as it is the header row")

        if ws.max_row > max_needed_row:
            logger_instance.info(f"Deleting {ws.max_row - max_needed_row} rows after row {max_needed_row}")
            ws.delete_rows(max_needed_row + 1, ws.max_row - max_needed_row)

        logger_instance.info("Setting worksheet view to A1")
        ws.sheet_view.topLeftCell = 'A1'

        # Save to ORIGINAL filename (not the cleaned one) so upload works correctly
        try:
            wb.save(original_filename)
            logger_instance.info(f"Excel file saved: {original_filename}, size: {os.path.getsize(original_filename)} bytes")
        finally:
            wb.close()
            logger_instance.info(f"Excel workbook closed: {original_filename}")

        # CRITICAL FIX: Fix openpyxl 3.1.5 bug with absolute image paths
        if not fix_excel_image_paths(original_filename, logger_instance):
            logger_instance.warning(f"Failed to fix image paths, file may be corrupted")

        # Clean up the temporary cleaned file
        if cleaned_filename != original_filename and os.path.exists(cleaned_filename):
            try:
                os.remove(cleaned_filename)
                logger_instance.info(f"Removed temporary cleaned file: {cleaned_filename}")
            except:
                pass

        # Validate the saved file
        logger_instance.info("Validating saved Excel file...")
        if not validate_excel_file(original_filename, logger_instance):
            logger_instance.error(f"Excel file failed validation after save: {original_filename}")
            raise ValueError(f"Generated Excel file is corrupted: {original_filename}")
    except Exception as e:
        logger_instance.error(f"Error writing to DISTRO Excel file: {e}", exc_info=True)
        raise


def clean_template_file(template_path: str, logger_instance: logging.Logger, mark_broken: bool = True) -> str:
    """
    Clean external links from template file by manually editing the ZIP archive.

    Args:
        template_path: Path to the template file
        logger_instance: Logger instance
        mark_broken: If True, keeps external links but marks them as broken (preserves formulas).
                     If False, removes external links entirely (old behavior).

    Returns path to cleaned file.
    """
    import zipfile
    import xml.etree.ElementTree as ET

    try:
        cleaned_path = f"{template_path}.cleaned.xlsx"
        logger_instance.info(f"Cleaning template file: {template_path} (mark_broken={mark_broken})")

        # Read the original Excel file as a ZIP
        with zipfile.ZipFile(template_path, 'r') as zip_in:
            file_list = zip_in.namelist()
            logger_instance.info(f"Original file contains {len(file_list)} entries")

            # Identify files that need cleaning
            external_files = [f for f in file_list if 'externalLink' in f.lower()]

            if external_files:
                logger_instance.info(f"Found external link files: {external_files}")

            # Create new ZIP
            with zipfile.ZipFile(cleaned_path, 'w', zipfile.ZIP_DEFLATED) as zip_out:
                for item in file_list:
                    # Handle external link files based on mode
                    if 'externalLink' in item.lower():
                        if mark_broken:
                            logger_instance.info(f"Keeping but marking as broken: {item}")
                            # Keep the file but we'll modify it below
                        else:
                            logger_instance.info(f"Removing external link file: {item}")
                            continue  # Skip this file entirely

                    # Get original file info to preserve metadata
                    info = zip_in.getinfo(item)
                    # Read the file content
                    data = zip_in.read(item)

                    # Modify external link files to point to dummy path (mark as broken)
                    if mark_broken and 'externalLinks/externalLink' in item and item.endswith('.xml'):
                        try:
                            ET.register_namespace('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')
                            root = ET.fromstring(data)

                            # Find externalBook element and change the link to a non-existent dummy path
                            for elem in root.iter():
                                if 'externalBook' in elem.tag:
                                    # Change r:id to point to a broken relationship
                                    # We'll handle this in the .rels file
                                    pass

                            # Convert back to bytes
                            data = ET.tostring(root, encoding='utf-8', xml_declaration=False)
                            data = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + data
                            logger_instance.info(f"Modified external link file: {item}")
                        except Exception as xml_error:
                            logger_instance.warning(f"Could not parse external link file {item}: {xml_error}, using original")

                    # Modify external link .rels files to point to non-existent paths
                    if mark_broken and 'externalLinks/_rels/externalLink' in item and item.endswith('.rels'):
                        try:
                            ET.register_namespace('', 'http://schemas.openxmlformats.org/package/2006/relationships')
                            root = ET.fromstring(data)

                            # Change all Target paths to a dummy non-existent file
                            for rel in root.iter():
                                if 'Relationship' in rel.tag:
                                    if rel.get('Target'):
                                        # Change to a clearly broken path
                                        rel.set('Target', 'file:///BROKEN_EXTERNAL_LINK.xlsx')
                                        rel.set('TargetMode', 'External')
                                        logger_instance.info(f"Marked relationship as broken in {item}")

                            data = ET.tostring(root, encoding='utf-8', xml_declaration=False)
                            data = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + data
                            logger_instance.info(f"Modified external link relationship: {item}")
                        except Exception as xml_error:
                            logger_instance.warning(f"Could not parse external link .rels {item}: {xml_error}, using original")

                    # Clean workbook.xml to remove externalReferences (only if NOT mark_broken)
                    if item == 'xl/workbook.xml' and not mark_broken:
                        try:
                            # Register ALL common namespaces to preserve XML structure
                            ET.register_namespace('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')
                            ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
                            ET.register_namespace('mc', 'http://schemas.openxmlformats.org/markup-compatibility/2006')
                            ET.register_namespace('x15', 'http://schemas.microsoft.com/office/spreadsheetml/2010/11/main')

                            root = ET.fromstring(data)

                            # Remove externalReferences element (with and without namespace)
                            # Try with namespace
                            for elem in root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}externalReferences'):
                                root.remove(elem)
                                logger_instance.info("Removed externalReferences from workbook.xml (with namespace)")

                            # Try without namespace (fallback)
                            for elem in root.findall('.//externalReferences'):
                                root.remove(elem)
                                logger_instance.info("Removed externalReferences from workbook.xml (without namespace)")

                            # Preserve the original XML declaration and structure
                            data = ET.tostring(root, encoding='utf-8', xml_declaration=False)
                            # Add proper XML declaration manually
                            data = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + data
                        except Exception as xml_error:
                            logger_instance.warning(f"Could not parse workbook.xml: {xml_error}, using original")

                    # Clean workbook.xml.rels to remove external link relationships (only if NOT mark_broken)
                    elif 'workbook.xml.rels' in item and not mark_broken:
                        try:
                            # Register namespace
                            ET.register_namespace('', 'http://schemas.openxmlformats.org/package/2006/relationships')

                            root = ET.fromstring(data)
                            ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}

                            # Remove Relationship elements that point to externalLinks
                            for rel in list(root.findall('.//r:Relationship', ns)):
                                target = rel.get('Target', '')
                                rel_type = rel.get('Type', '')
                                if 'externalLink' in target or 'externalLink' in rel_type:
                                    rel_id = rel.get('Id')
                                    logger_instance.info(f"Removing external link relationship: {rel_id} -> {target}")
                                    root.remove(rel)

                            # Preserve the original XML declaration
                            data = ET.tostring(root, encoding='utf-8', xml_declaration=False)
                            data = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + data
                        except Exception as xml_error:
                            logger_instance.warning(f"Could not parse workbook.xml.rels: {xml_error}, using original")

                    # Write to new ZIP, preserving original compression metadata
                    zip_out.writestr(info, data)

        logger_instance.info(f"Created cleaned template: {cleaned_path}")

        # Verify the cleaned file can be loaded (without data_only to ensure full functionality)
        try:
            test_wb = load_workbook(cleaned_path)
            test_wb.close()
            logger_instance.info("Verified cleaned template can be loaded successfully")
            return cleaned_path
        except Exception as verify_error:
            logger_instance.error(f"Cleaned file verification failed: {verify_error}")
            logger_instance.info("Attempting to use original file with data_only=True verification...")
            # Try to verify original can at least be loaded with data_only
            try:
                test_wb2 = load_workbook(template_path, data_only=True)
                test_wb2.close()
                logger_instance.warning("Original file can only be loaded with data_only=True - images may not work properly")
                return template_path
            except:
                logger_instance.error("Both cleaned and original files have issues")
                return template_path

    except Exception as e:
        logger_instance.error(f"Failed to clean template file: {e}", exc_info=True)
        return template_path  # Return original if cleaning fails

def validate_excel_file(excel_file: str, logger_instance: logging.Logger) -> bool:
    """Validate that the Excel file can be opened without corruption."""
    wb = None
    try:
        logger_instance.info(f"Starting Excel file validation for: {excel_file}")

        # Check file exists and has size
        if not os.path.exists(excel_file):
            logger_instance.error(f"Excel file does not exist: {excel_file}")
            return False

        file_size = os.path.getsize(excel_file)
        logger_instance.info(f"Excel file size: {file_size} bytes")

        if file_size < 5000:  # Excel files should be at least 5KB
            logger_instance.error(f"Excel file too small: {file_size} bytes")
            return False

        # Try to open the file
        wb = load_workbook(excel_file, data_only=False)
        ws = wb.active

        # Try to access some basic properties
        max_row = ws.max_row
        max_col = ws.max_column
        logger_instance.info(f"Excel file dimensions: {max_row} rows x {max_col} columns")

        # Check if there are any images
        if hasattr(ws, '_images'):
            image_count = len(ws._images)
            logger_instance.info(f"Excel file contains {image_count} images")

            # Validate each image
            for idx, img in enumerate(ws._images):
                try:
                    img_width = getattr(img, 'width', 0)
                    img_height = getattr(img, 'height', 0)
                    anchor = getattr(img, 'anchor', None)
                    anchor_type = type(anchor).__name__ if anchor else 'None'

                    # Get detailed anchor info
                    anchor_details = ""
                    if anchor and hasattr(anchor, '_from'):
                        marker = anchor._from
                        anchor_details = f", pos(col={getattr(marker, 'col', '?')}, row={getattr(marker, 'row', '?')})"
                        if hasattr(marker, 'colOff') and hasattr(marker, 'rowOff'):
                            anchor_details += f", off({getattr(marker, 'colOff', 0)}, {getattr(marker, 'rowOff', 0)})"

                    logger_instance.info(f"  Image {idx+1}: {img_width}x{img_height}, anchor={anchor_type}{anchor_details}")
                except Exception as img_check_error:
                    logger_instance.warning(f"  Image {idx+1} validation warning: {img_check_error}")

        logger_instance.info(f"Excel file validation PASSED: {excel_file}")
        return True
    except Exception as e:
        logger_instance.error(f"Excel file validation FAILED: {excel_file}, error: {e}", exc_info=True)
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except:
                pass

def find_header_row_index(excel_file: str, logger_instance: logging.Logger) -> Optional[int]:
    wb = None
    try:
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
        keywords = {'id', 'name', 'product', 'brand', 'sku', 'category', 'color', 'price', 'description'}
        best_row, best_score = None, 0
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if not row:
                continue
            cells = [str(c).strip().lower() for c in row if isinstance(c, str)]
            score = len(cells) + len(set(cells)) + sum(any(k in c for k in keywords) for c in cells)
            if score > best_score:
                best_score, best_row = score, i
        if best_row and best_score > 3:
            return best_row - 1  # Convert to 0-based indexing
        return None
    except Exception as e:
        logger_instance.error(f"Could not find header row: {e}")
        return None
    finally:
        if wb is not None:
            try:
                wb.close()
            except:
                pass

def write_excel_msrp(local_filename: str, temp_dir: str, image_data: List[Dict], header_row: int, target_column: str, row_offset: int, logger_instance: logging.Logger, populate_images: bool = True, populate_msrp: bool = True):
    try:
        # Clean the file (mark external links as broken to preserve formulas)
        logger_instance.info("Cleaning file (marking external links as broken)...")
        original_filename = local_filename
        cleaned_filename = clean_template_file(local_filename, logger_instance, mark_broken=True)

        # Load the workbook
        wb = load_workbook(cleaned_filename)
        ws = wb.active

        # CRITICAL FIX: Clear ALL existing images to prevent TwoCellAnchor/OneCellAnchor conflicts
        if hasattr(ws, '_images') and ws._images:
            existing_count = len(ws._images)
            ws._images = []
            logger_instance.info(f"Cleared {existing_count} existing images from template to prevent anchor conflicts")

        image_map = {int(Path(f).stem): f for f in os.listdir(temp_dir) if Path(f).stem.isdigit()}
        if populate_msrp and not re.match(r'^[A-Z]+$', target_column):
            raise ValueError(f"Invalid target_column: {target_column}. Must be a valid Excel column letter (e.g., 'A', 'B', 'AA').")

        header_row = int(header_row)
        if header_row < 0:
            logger_instance.error(f"Invalid header_row {header_row}. Setting to 0.")
            header_row = 0
        if row_offset < 0:
            logger_instance.warning(f"Negative row_offset {row_offset} provided. Ensure this is intentional.")

        row_dim = ws.row_dimensions.get(header_row + 1)
        DEFAULT_ROW_HEIGHT_POINTS = row_dim.height if row_dim and row_dim.height is not None else 12.75
        logger_instance.info(f"Using template row height: {DEFAULT_ROW_HEIGHT_POINTS} points from row {header_row + 1}")

        row_data_map = {item['ExcelRowID']: item for item in image_data}
        last_non_empty_row = get_last_non_empty_row(ws, column='B', header_row=header_row, logger_instance=logger_instance)

        data_row_ids = sorted(row_data_map.keys())
        if data_row_ids:
            min_row_id = data_row_ids[0]
            max_row_id = data_row_ids[-1]
        else:
            min_row_id = 1
            relative_last_row = last_non_empty_row - (header_row + 1)
            relative_last_row = max(1, relative_last_row)
            max_row_id = relative_last_row
            logger_instance.warning(
                f"No data in image_data, deriving row range 1-{max_row_id} from template last row {last_non_empty_row}"
            )

        base_row = header_row + row_offset + 2
        if base_row < 1:
            logger_instance.error(f"Computed base_row {base_row} is less than 1. Aborting.")
            raise ValueError("Invalid row range: base_row is less than 1")

        # CRITICAL FIX: Only create rows that have actual data
        row_id_to_row_num = {}
        for idx, row_id in enumerate(data_row_ids):
            row_id_to_row_num[row_id] = base_row + idx

        max_needed_row = base_row + len(data_row_ids) - 1 if data_row_ids else base_row
        max_needed_row = max(max_needed_row, last_non_empty_row)

        logger_instance.info(
            f"Row mapping summary -> base_row: {base_row}, min_row_id: {min_row_id}, max_row_id: {max_row_id}, "
            f"data_rows_count: {len(data_row_ids)}, max_needed_row: {max_needed_row}"
        )

        if ws.max_row < max_needed_row:
            logger_instance.info(f"Appending {max_needed_row - ws.max_row} rows to worksheet")
            for row_num in range(ws.max_row + 1, max_needed_row + 1):
                ws.append([''] * ws.max_column)
                ws.row_dimensions[row_num].height = DEFAULT_ROW_HEIGHT_POINTS

        for row_id in data_row_ids:
            row_num = row_id_to_row_num[row_id]

            ws.row_dimensions[row_num].height = DEFAULT_ROW_HEIGHT_POINTS

            if row_id in row_data_map:
                item = row_data_map[row_id]
                if populate_images:
                    if row_id in image_map:
                        image_path = os.path.join(temp_dir, image_map[row_id])
                        if verify_and_process_image(image_path, logger_instance):
                            try:
                                # Validate image file before inserting
                                if not os.path.exists(image_path) or os.path.getsize(image_path) < MIN_IMAGE_SIZE:
                                    logger_instance.warning(f"Image file invalid after processing for Row {row_id}, skipping")
                                else:
                                    # CRITICAL: Validate PIL can read the image BEFORE creating openpyxl Image
                                    try:
                                        pil_img = PILImage.open(image_path)
                                        pil_width, pil_height = pil_img.size
                                        pil_img.close()
                                        if pil_width <= 0 or pil_height <= 0:
                                            logger_instance.warning(f"PIL image has invalid dimensions for Row {row_id}: {pil_width}x{pil_height}, skipping")
                                            continue
                                    except Exception as pil_error:
                                        logger_instance.warning(f"PIL failed to read image for Row {row_id}: {pil_error}, skipping")
                                        continue

                                    img = Image(image_path)

                                    img_height_pixels = img.height if hasattr(img, 'height') else 0
                                    img_width_pixels = img.width if hasattr(img, 'width') else 0

                                    # Validate dimensions
                                    if img_height_pixels <= 0 or img_width_pixels <= 0:
                                        logger_instance.warning(f"Invalid image dimensions for Row {row_id}: {img_width_pixels}x{img_height_pixels}, skipping")
                                    else:
                                        img_height_points = img_height_pixels * 72 / 96

                                        new_height = max(DEFAULT_ROW_HEIGHT_POINTS, img_height_points)
                                        ws.row_dimensions[row_num].height = new_height

                                        # Calculate dimensions for centering
                                        col_width_chars = ws.column_dimensions['A'].width
                                        if col_width_chars is None:
                                            col_width_chars = 8.43
                                        cell_width_pixels = col_width_chars * 7
                                        cell_height_pixels = points_to_pixels(new_height)

                                        x_offset_pixels = max(0, (cell_width_pixels - img_width_pixels) / 2)
                                        y_offset_pixels = max(0, (cell_height_pixels - img_height_pixels) / 2)

                                        width_emu = pixels_to_EMU(img_width_pixels)
                                        height_emu = pixels_to_EMU(img_height_pixels)

                                        # Validate EMU values to prevent XML corruption
                                        if width_emu <= 0 or height_emu <= 0:
                                            logger_instance.warning(f"Invalid EMU dimensions for Row {row_id}: {width_emu}x{height_emu}, skipping")
                                        elif width_emu > 10000000 or height_emu > 10000000:
                                            logger_instance.warning(f"EMU dimensions too large for Row {row_id}: {width_emu}x{height_emu}, skipping")
                                        else:
                                            anchor_row = max(0, row_num - 1)
                                            # Ensure offsets are within valid range (max ~1000 pixels)
                                            MAX_OFFSET_EMU = 9525000
                                            x_offset_emu = max(0, min(pixels_to_EMU(x_offset_pixels), MAX_OFFSET_EMU))
                                            y_offset_emu = max(0, min(pixels_to_EMU(y_offset_pixels), MAX_OFFSET_EMU))

                                            marker = AnchorMarker(
                                                col=0, # Column A
                                                colOff=x_offset_emu,
                                                row=anchor_row,
                                                rowOff=y_offset_emu
                                            )
                                            img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(width_emu, height_emu))

                                            # CRITICAL: Final validation before adding to worksheet
                                            final_width = getattr(img, 'width', 0)
                                            final_height = getattr(img, 'height', 0)
                                            if final_width <= 0 or final_height <= 0:
                                                logger_instance.error(f"CRITICAL: Image dimensions became 0 after anchor was set for Row {row_id}: {final_width}x{final_height}")
                                                logger_instance.warning(f"Skipping image insertion for Row {row_id}")
                                            else:
                                                ws.add_image(img)

                                        logger_instance.info(f"Added image for Row {row_id} at Excel row {row_num}, height set to {ws.row_dimensions[row_num].height} points")
                            except Exception as img_error:
                                logger_instance.error(f"Failed to insert image for Row {row_id}: {img_error}", exc_info=True)
                                logger_instance.warning(f"Skipping image insertion for Row {row_id}, writing MSRP only")
                        else:
                            logger_instance.warning(f"Image processing failed for Row {row_id}, writing MSRP only")
                    else:
                        logger_instance.info(f"No image found for Row {row_id}, writing MSRP only")

                if populate_msrp:
                    msrp_value = item.get('MSRP', '')
                    ws.cell(row=row_num, column=column_index_from_string(target_column)).value = msrp_value
                    logger_instance.info(f"Wrote MSRP '{msrp_value}' for Row {row_id} at {target_column}{row_num}")
            else:
                if populate_msrp:
                    ws.cell(row=row_num, column=column_index_from_string(target_column)).value = ''
                    logger_instance.info(f"Filled missing row {row_num} (ExcelRowID {row_id}) with empty MSRP")

        if ws.max_row > max_needed_row:
            logger_instance.info(f"Deleting {ws.max_row - max_needed_row} rows after row {max_needed_row}")
            ws.delete_rows(max_needed_row + 1, ws.max_row - max_needed_row)

        logger_instance.info("Setting worksheet view to A1")
        ws.sheet_view.topLeftCell = 'A1'

        # Save to ORIGINAL filename (not the cleaned one) so upload works correctly
        try:
            wb.save(original_filename)
            logger_instance.info(f"MSRP Excel file saved: {original_filename}, size: {os.path.getsize(original_filename)} bytes")
        finally:
            wb.close()
            logger_instance.info(f"Excel workbook closed: {original_filename}")

        # CRITICAL FIX: Fix openpyxl 3.1.5 bug with absolute image paths
        if not fix_excel_image_paths(original_filename, logger_instance):
            logger_instance.warning(f"Failed to fix image paths, file may be corrupted")

        # Clean up the temporary cleaned file
        if cleaned_filename != original_filename and os.path.exists(cleaned_filename):
            try:
                os.remove(cleaned_filename)
                logger_instance.info(f"Removed temporary cleaned file: {cleaned_filename}")
            except:
                pass

        # Validate the saved file
        if not validate_excel_file(original_filename, logger_instance):
            logger_instance.error(f"Excel file failed validation after save: {original_filename}")
            raise ValueError(f"Generated Excel file is corrupted: {original_filename}")
    except Exception as e:
        logger_instance.error(f"Error writing to MSRP Excel file: {e}", exc_info=True)
        raise

def write_excel_generic(local_filename: str, temp_dir: str, image_data: List[Dict], header_row: int, row_offset: int, logger_instance: logging.Logger, file_type_id: Optional[int] = None):
    try:
        # Clean the file (mark external links as broken to preserve formulas)
        logger_instance.info("Cleaning file (marking external links as broken)...")
        original_filename = local_filename
        cleaned_filename = clean_template_file(local_filename, logger_instance, mark_broken=True)

        # Load the workbook
        wb = load_workbook(cleaned_filename)
        ws = wb.active

        # CRITICAL FIX: Clear ALL existing images to prevent TwoCellAnchor/OneCellAnchor conflicts
        # This prevents drawing XML corruption when mixing anchor types
        if hasattr(ws, '_images') and ws._images:
            existing_count = len(ws._images)
            ws._images = []
            logger_instance.info(f"Cleared {existing_count} existing images from template to prevent anchor conflicts")

        temp_path = Path(temp_dir)
        image_map = {}
        for path_obj in temp_path.iterdir():
            if not path_obj.is_file():
                continue
            stem = path_obj.stem
            if stem.isdigit():
                image_map[int(stem)] = path_obj

        row_ids_from_data = set()
        for item in image_data:
            value = item.get('ExcelRowID')
            if value is None:
                continue
            try:
                row_ids_from_data.add(int(value))
            except (ValueError, TypeError):
                logger_instance.warning(f"Skipping invalid ExcelRowID '{value}' in generic writer")

        all_row_ids = sorted(row_ids_from_data.union(image_map.keys()))

        if not all_row_ids:
            logger_instance.info("No rows found to write for generic template.")
            return

        min_row_id = all_row_ids[0]
        max_row_id = all_row_ids[-1]

        template_row_dim = ws.row_dimensions.get(header_row + 1)
        default_row_height = template_row_dim.height if template_row_dim and template_row_dim.height is not None else 12.75

        header_excel_row = max(header_row, 0) + 1
        use_absolute_row_ids = min_row_id > header_excel_row
        if use_absolute_row_ids:
            logger_instance.info(
                f"Detected absolute ExcelRowID values (min_row_id={min_row_id}, header_excel_row={header_excel_row})."
            )
        else:
            logger_instance.info(
                f"Using relative ExcelRowID mapping (min_row_id={min_row_id}, header_excel_row={header_excel_row})."
            )

        row_num_map: Dict[int, int] = {}
        total_columns = max(ws.max_column, 1)
        relative_base_row = header_row + row_offset + 2
        for row_id in all_row_ids:
            if use_absolute_row_ids:
                row_num = row_id + row_offset
            else:
                row_num = relative_base_row + (row_id - min_row_id)

            if row_num < 1:
                logger_instance.error(
                    f"Computed row {row_num} is invalid for row_id={row_id}, header_row={header_row}, row_offset={row_offset}. Skipping."
                )
                continue
            row_num_map[row_id] = row_num

        if not row_num_map:
            logger_instance.warning("No valid row mappings computed for generic template. Nothing to write.")
            return

        max_needed_row = max(row_num_map.values())
        if ws.max_row < max_needed_row:
            for _ in range(ws.max_row + 1, max_needed_row + 1):
                ws.append([''] * total_columns)

        for row_id in all_row_ids:
            row_num = row_num_map.get(row_id)
            if row_num is None:
                continue

            image_path_obj = image_map.get(row_id)
            if not image_path_obj:
                continue

            image_path = str(image_path_obj)
            if verify_and_process_image(image_path, logger_instance):
                try:
                    # Validate image file before inserting
                    if not os.path.exists(image_path) or os.path.getsize(image_path) < MIN_IMAGE_SIZE:
                        logger_instance.warning(f"Image file invalid after processing for ExcelRowID {row_id}, skipping")
                    else:
                        # CRITICAL: Validate PIL can read the image BEFORE creating openpyxl Image
                        try:
                            pil_img = PILImage.open(image_path)
                            pil_width, pil_height = pil_img.size
                            pil_img.close()
                            if pil_width <= 0 or pil_height <= 0:
                                logger_instance.warning(f"PIL image has invalid dimensions for ExcelRowID {row_id}: {pil_width}x{pil_height}, skipping")
                                continue
                        except Exception as pil_error:
                            logger_instance.warning(f"PIL failed to read image for ExcelRowID {row_id}: {pil_error}, skipping")
                            continue

                        img = Image(image_path)

                        img_height_pixels = getattr(img, 'height', 0)
                        img_width_pixels = getattr(img, 'width', 0)

                        # Validate dimensions
                        if img_height_pixels <= 0 or img_width_pixels <= 0:
                            logger_instance.warning(f"Invalid image dimensions for ExcelRowID {row_id}: {img_width_pixels}x{img_height_pixels}, skipping")
                        else:
                            img_height_points = img_height_pixels * 72 / 96
                            current_height = ws.row_dimensions[row_num].height
                            new_height = max(img_height_points, current_height or default_row_height)
                            ws.row_dimensions[row_num].height = new_height

                            # Calculate dimensions for centering
                            col_width_chars = ws.column_dimensions['A'].width
                            if col_width_chars is None:
                                col_width_chars = 8.43
                            cell_width_pixels = col_width_chars * 7
                            cell_height_pixels = points_to_pixels(new_height)

                            x_offset_pixels = max(0, (cell_width_pixels - img_width_pixels) / 2)
                            y_offset_pixels = max(0, (cell_height_pixels - img_height_pixels) / 2)

                            width_emu = pixels_to_EMU(img_width_pixels)
                            height_emu = pixels_to_EMU(img_height_pixels)

                            # Validate EMU values to prevent XML corruption
                            if width_emu <= 0 or height_emu <= 0:
                                logger_instance.warning(f"Invalid EMU dimensions for ExcelRowID {row_id}: {width_emu}x{height_emu}, skipping")
                            elif width_emu > 10000000 or height_emu > 10000000:
                                logger_instance.warning(f"EMU dimensions too large for ExcelRowID {row_id}: {width_emu}x{height_emu}, skipping")
                            else:
                                anchor_row = max(0, row_num - 1)
                                # Ensure offsets are within valid range (max ~1000 pixels)
                                MAX_OFFSET_EMU = 9525000
                                x_offset_emu = max(0, min(pixels_to_EMU(x_offset_pixels), MAX_OFFSET_EMU))
                                y_offset_emu = max(0, min(pixels_to_EMU(y_offset_pixels), MAX_OFFSET_EMU))

                                marker = AnchorMarker(
                                    col=0, # Column A
                                    colOff=x_offset_emu,
                                    row=anchor_row,
                                    rowOff=y_offset_emu
                                )
                                img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(width_emu, height_emu))

                                # CRITICAL: Final validation before adding to worksheet
                                final_width = getattr(img, 'width', 0)
                                final_height = getattr(img, 'height', 0)
                                if final_width <= 0 or final_height <= 0:
                                    logger_instance.error(f"CRITICAL: Image dimensions became 0 after anchor was set for ExcelRowID {row_id}: {final_width}x{final_height}")
                                    logger_instance.warning(f"Skipping image insertion for ExcelRowID {row_id}")
                                else:
                                    ws.add_image(img)

                            logger_instance.info(
                                f"Added image for ExcelRowID {row_id} at Excel row {row_num} (absolute_mapping={use_absolute_row_ids})"
                            )
                except Exception as img_error:
                    logger_instance.error(f"Failed to insert image for ExcelRowID {row_id}: {img_error}", exc_info=True)
                    logger_instance.warning(f"Skipping image insertion for ExcelRowID {row_id}")
            else:
                logger_instance.warning(f"Image verification failed for ExcelRowID {row_id} at path {image_path}")

        logger_instance.info("Setting worksheet view to A1")
        ws.sheet_view.topLeftCell = 'A1'

        # Save to ORIGINAL filename (not the cleaned one) so upload works correctly
        try:
            wb.save(original_filename)
            logger_instance.info(f"Generic Excel file saved: {original_filename}, size: {os.path.getsize(original_filename)} bytes")
        finally:
            wb.close()
            logger_instance.info(f"Excel workbook closed: {original_filename}")

        # CRITICAL FIX: Fix openpyxl 3.1.5 bug with absolute image paths
        if not fix_excel_image_paths(original_filename, logger_instance):
            logger_instance.warning(f"Failed to fix image paths, file may be corrupted")

        # Clean up the temporary cleaned file
        if cleaned_filename != original_filename and os.path.exists(cleaned_filename):
            try:
                os.remove(cleaned_filename)
                logger_instance.info(f"Removed temporary cleaned file: {cleaned_filename}")
            except:
                pass

        # Validate the saved file
        if not validate_excel_file(original_filename, logger_instance):
            logger_instance.error(f"Excel file failed validation after save: {original_filename}")
            raise ValueError(f"Generated Excel file is corrupted: {original_filename}")
    except Exception as e:
        logger_instance.error(f"Error writing to generic Excel file: {e}", exc_info=True)
        raise

# --- Main Background Task ---
async def generate_download_file(file_id: str, row_offset: int = 0):
    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    logger_instance = setup_logging(file_id, timestamp)
    temp_images_dir, temp_excel_dir = "", ""
    file_id_int = int(file_id)

    try:
        user_agents = get_user_agents(logger_instance)
        temp_images_dir, temp_excel_dir = await create_temp_dirs(file_id)
        
        logger_instance.info("Fetching all image data from database...")
        images_df = get_images_excel_db(file_id_int, logger_instance)
        
        if images_df.empty:
            logger_instance.warning(f"No image data found for FileID {file_id}. Fetching all records for metadata.")
            query = """
                SELECT
                    s.ExcelRowID,
                    s.ProductBrand AS Brand, s.ProductModel AS Style,
                    s.ProductColor AS Color, s.ProductCategory AS Category
                FROM utb_ImageScraperFiles f
                INNER JOIN utb_ImageScraperRecords s ON s.FileID = f.ID 
                WHERE f.ID = ?
                ORDER BY s.ExcelRowID
            """
            images_df = pd.read_sql_query(query, engine, params=[(file_id,)])
            if images_df.empty:
                logger_instance.error(f"No records found for FileID {file_id}. The job cannot proceed.")
                return

        logger_instance.info(f"ExcelRowID values: {list(images_df['ExcelRowID'])}")
        logger_instance.info("Grouping data by product to prepare for download attempts...")
        grouped_data = []
        for _, group in images_df.groupby('ExcelRowID'):
            first_row = group.iloc[0]
            image_options = list(zip(group['ImageUrl'], group['ImageUrlThumbnail'], group['SortOrder'])) if 'ImageUrl' in group.columns else []
            
            grouped_data.append({
                'ExcelRowID': int(first_row['ExcelRowID']),
                'Brand': first_row['Brand'],
                'Style': first_row['Style'],
                'Color': first_row['Color'],
                'Category': first_row['Category'],
                'image_options': image_options
            })
        
        logger_instance.info(f"Data prepared for {len(grouped_data)} unique products. Starting image downloads.")
        if any(item['image_options'] for item in grouped_data):
            await download_all_images(grouped_data, temp_images_dir, logger_instance, user_agents)
        else:
            logger_instance.info("No images to download, proceeding with metadata only.")
        
        file_type_id = get_file_type_id(file_id_int, logger_instance)
        FILE_TYPE_DISTRO = 3

        file_url, header_row_from_db, user_email = get_file_location_and_header(file_id_int, logger_instance)
        original_file_name = os.path.basename(urllib.parse.unquote(file_url))

        if file_type_id == FILE_TYPE_DISTRO:
            logger_instance.info("Starting DISTRO file generation.")
            template_url = "https://iconluxury.shop/documents/public_ICON_DISTRO_USD_20250617.xlsx"
            file_name = os.path.basename(urllib.parse.unquote(template_url))
            local_filename = os.path.join(temp_excel_dir, file_name)
            header_row = 4  # Hardcoded for DISTRO (0-based)

            if row_offset:
                logger_instance.info(
                    f"Ignoring provided row_offset={row_offset} for DISTRO template; using fixed layout."
                )
            template_row_offset = 0

            res = requests.get(template_url, timeout=60)
            res.raise_for_status()
            with open(local_filename, "wb") as f:
                f.write(res.content)
            write_excel_distro(local_filename, temp_images_dir, grouped_data, header_row, logger_instance, template_row_offset)
        else:
            logger_instance.info("Starting GENERIC file generation.")
            file_name = os.path.basename(urllib.parse.unquote(file_url))
            local_filename = os.path.join(temp_excel_dir, file_name)

            res = requests.get(file_url, timeout=60)
            res.raise_for_status()
            with open(local_filename, "wb") as f:
                f.write(res.content)

            inferred_header_row = find_header_row_index(local_filename, logger_instance)
            if header_row_from_db is not None:
                try:
                    db_header = int(header_row_from_db)
                except (TypeError, ValueError):
                    logger_instance.warning(f"Invalid header_row_from_db '{header_row_from_db}'. Falling back to inferred value.")
                    db_header = None

                if db_header is not None:
                    candidates = {max(0, db_header - 1), max(0, db_header)}
                    if inferred_header_row is not None:
                        header_row_value = min(candidates, key=lambda c: abs(c - inferred_header_row))
                    else:
                        header_row_value = min(candidates)
                else:
                    header_row_value = inferred_header_row if inferred_header_row is not None else 0
            else:
                header_row_value = inferred_header_row if inferred_header_row is not None else 0

            logger_instance.info(f"Using header_row_value={header_row_value} for generic template.")
            write_excel_generic(local_filename, temp_images_dir, grouped_data, header_row_value, row_offset, logger_instance, file_type_id=file_type_id)
            
        processed_file_name = f"{Path(original_file_name).stem}_ptype-{file_type_id}_{timestamp}.xlsx"
        public_url = await upload_file_to_space(local_filename, save_as=f"processed_files/{processed_file_name}", file_id=file_id_int, is_public=True)
        update_file_location_complete(file_id_int, public_url, logger_instance)
        await send_email(
            to_emails=user_email or '',
            subject=f'File Processed: {file_name}',
            file_path=local_filename,
            job_id=file_id
        )
        
        logger_instance.info(f"Successfully completed job for FileID {file_id}.")

    except Exception as e:
        logger_instance.error(f"FATAL ERROR for FileID {file_id}: {e}", exc_info=True)
    finally:
        if temp_images_dir and temp_excel_dir:
            await cleanup_temp_dirs([temp_images_dir, temp_excel_dir])

async def generate_msrp_excel(file_id: str, target_column: str, row_offset: int = 0):
    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    logger_instance = setup_logging(file_id, timestamp)
    temp_images_dir, temp_excel_dir = "", ""
    file_id_int = int(file_id)

    try:
        user_agents = get_user_agents(logger_instance)
        temp_images_dir, temp_excel_dir = await create_temp_dirs(file_id)
        
        # Determine what to populate based on FileTypeID
        file_type_id = get_file_type_id(file_id_int, logger_instance)
        populate_images = True
        populate_msrp = True
        
        if file_type_id == 8:
            populate_msrp = False
            logger_instance.info(f"FileTypeID is 8: Populating IMAGES ONLY.")
        elif file_type_id == 9:
            populate_images = False
            logger_instance.info(f"FileTypeID is 9: Populating MSRP ONLY.")

        logger_instance.info("Fetching all image and MSRP data from database...")
        images_df = get_images_excel_db(file_id_int, logger_instance)
        
        if images_df.empty:
            logger_instance.warning(f"No data found for FileID {file_id}. Fetching all records for MSRP.")
            query = """
                SELECT
                    s.ExcelRowID,
                    s.ProductMSRP AS MSRP
                FROM utb_ImageScraperFiles f
                INNER JOIN utb_ImageScraperRecords s ON s.FileID = f.ID 
                WHERE f.ID = ?
                ORDER BY s.ExcelRowID
            """
            images_df = pd.read_sql_query(query, engine, params=[(file_id,)])
            if images_df.empty:
                logger_instance.error(f"No records found for FileID {file_id}. The job cannot proceed.")
                return

        logger_instance.info("Grouping data by product to prepare for download attempts...")
        grouped_data = []
        for _, group in images_df.groupby('ExcelRowID'):
            first_row = group.iloc[0]
            image_options = list(zip(group['ImageUrl'], group['ImageUrlThumbnail'], group['SortOrder'])) if 'ImageUrl' in group.columns else []
            
            grouped_data.append({
                'ExcelRowID': int(first_row['ExcelRowID']),
                'MSRP': first_row.get('MSRP', ''),
                'image_options': image_options
            })
        
        logger_instance.info(f"Data prepared for {len(grouped_data)} unique products. Starting image downloads.")
        
        if populate_images:
            if any(item['image_options'] for item in grouped_data):
                await download_all_images(grouped_data, temp_images_dir, logger_instance, user_agents)
            else:
                logger_instance.info("No images to download.")
        else:
            logger_instance.info("Skipping image download as populate_images is False.")
        
        logger_instance.info("Starting MSRP file generation.")
        file_url, header_row_from_db, user_email = get_file_location_and_header(file_id_int, logger_instance)
        file_name = os.path.basename(urllib.parse.unquote(file_url))
        local_filename = os.path.join(temp_excel_dir, file_name)

        res = requests.get(file_url, timeout=60)
        res.raise_for_status()
        with open(local_filename, "wb") as f:
            f.write(res.content)

        inferred_header_row = find_header_row_index(local_filename, logger_instance)
        if header_row_from_db is not None:
            try:
                db_header = int(header_row_from_db)
            except (TypeError, ValueError):
                logger_instance.warning(f"Invalid header_row_from_db '{header_row_from_db}'. Falling back to inferred value.")
                db_header = None

            if db_header is not None:
                candidates = {max(0, db_header - 1), max(0, db_header)}
                if inferred_header_row is not None:
                    header_row_value = min(candidates, key=lambda c: abs(c - inferred_header_row))
                else:
                    header_row_value = min(candidates)
            else:
                header_row_value = inferred_header_row if inferred_header_row is not None else 0
        else:
            header_row_value = inferred_header_row if inferred_header_row is not None else 0

        logger_instance.info(f"Using header_row_value={header_row_value} for MSRP template.")
        write_excel_msrp(local_filename, temp_images_dir, grouped_data, header_row_value, target_column, row_offset, logger_instance, populate_images=populate_images, populate_msrp=populate_msrp)

        processed_file_name = f"{Path(file_name).stem}_msrp_{file_type_id}_{timestamp}.xlsx"
        public_url = await upload_file_to_space(local_filename, save_as=f"processed_files/{processed_file_name}", file_id=file_id_int, is_public=True)
        update_file_location_complete(file_id_int, public_url, logger_instance)
        await send_email(
            to_emails=user_email or 'nik@iconluxurygroup.com',
            subject=f'MSRP File Processed: {file_name}',
            file_path=local_filename,
            job_id=file_id
        )
        
        logger_instance.info(f"Successfully completed MSRP job for FileID {file_id}.")

    except Exception as e:
        logger_instance.error(f"FATAL ERROR for MSRP FileID {file_id}: {e}", exc_info=True)
    finally:
        if temp_images_dir and temp_excel_dir:
            await cleanup_temp_dirs([temp_images_dir, temp_excel_dir])

# --- FastAPI Endpoints ---
@app.post("/generate-download-file/")
async def process_file(background_tasks: BackgroundTasks, file_id: int, row_offset: Optional[int] = 0):
    logger.info(f"Received request for FileID: {file_id} with row_offset={row_offset}")
    background_tasks.add_task(generate_download_file, str(file_id), row_offset)
    return {"message": "Processing started. You will be notified upon completion."}

@app.post("/generate-msrp-excel/")
async def process_msrp_file(background_tasks: BackgroundTasks, file_id: int, target_column: str, row_offset: Optional[int] = 0):
    logger.info(f"Received request for MSRP FileID: {file_id} with target_column={target_column} and row_offset={row_offset}")
    background_tasks.add_task(generate_msrp_excel, str(file_id), target_column, row_offset)
    return {"message": "MSRP Processing started. You will be notified upon completion."}

@app.get("/")
def read_root():
    return {"message": "Image Scraper Service is running."}

if __name__ == "__main__":
    import uvicorn
    logger.info("Starting Uvicorn server on http://0.0.0.0:8080")
    uvicorn.run(app, host="0.0.0.0", port=8080)
