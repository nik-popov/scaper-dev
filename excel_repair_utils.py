"""
Utility module for repairing corrupted Excel files with external link issues.
This module provides functions to automatically detect and repair Excel files
before processing them with openpyxl.
"""

import logging
import os
import zipfile
import shutil
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Optional
from openpyxl import load_workbook


def repair_excel_file(input_file: str, output_file: Optional[str] = None, logger: Optional[logging.Logger] = None) -> str:
    """
    Repair an Excel file by removing corrupted external links and linked data types.

    Args:
        input_file: Path to the potentially corrupted Excel file
        output_file: Path for the repaired file (optional, defaults to input_file with _repaired suffix)
        logger: Optional logger instance for logging operations

    Returns:
        str: Path to the repaired file (or original if no repair was needed)
    """
    if logger is None:
        logger = logging.getLogger(__name__)

    input_path = Path(input_file)

    if output_file is None:
        output_file = str(input_path.parent / f"{input_path.stem}_repaired{input_path.suffix}")

    # Create a temporary directory for extraction
    temp_dir = input_path.parent / f"temp_excel_repair_{input_path.stem}"
    temp_dir.mkdir(exist_ok=True)

    try:
        logger.info(f"Checking Excel file for corruption: {input_file}")

        # Extract the Excel file (which is a ZIP archive)
        with zipfile.ZipFile(input_file, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)

        needs_repair = False

        # Check if external links exist
        external_links_dir = temp_dir / "xl" / "externalLinks"
        if external_links_dir.exists():
            logger.info(f"Found external links directory - removing...")
            shutil.rmtree(external_links_dir)
            needs_repair = True

        # Remove external link references from workbook.xml.rels
        rels_file = temp_dir / "xl" / "_rels" / "workbook.xml.rels"
        if rels_file.exists():
            with open(rels_file, 'r', encoding='utf-8') as f:
                content = f.read()

            if 'externalLinks' in content:
                logger.info(f"Cleaning external link relationships...")
                lines = content.split('\n')
                cleaned_lines = [line for line in lines if 'externalLinks' not in line]
                content = '\n'.join(cleaned_lines)

                with open(rels_file, 'w', encoding='utf-8') as f:
                    f.write(content)
                needs_repair = True

        # Remove external link references from [Content_Types].xml
        content_types_file = temp_dir / "[Content_Types].xml"
        if content_types_file.exists():
            with open(content_types_file, 'r', encoding='utf-8') as f:
                content = f.read()

            if 'externalLink' in content:
                logger.info(f"Cleaning content types...")
                lines = content.split('\n')
                cleaned_lines = [line for line in lines if 'externalLinks' not in line and 'externalLink' not in line]
                content = '\n'.join(cleaned_lines)

                with open(content_types_file, 'w', encoding='utf-8') as f:
                    f.write(content)
                needs_repair = True

        # Replace external formulas with #REF text in all worksheets
        worksheets_dir = temp_dir / "xl" / "worksheets"
        if worksheets_dir.exists():
            for sheet_file in worksheets_dir.glob("*.xml"):
                if _replace_external_formulas_with_text(sheet_file, logger):
                    needs_repair = True

        if needs_repair:
            # Create the repaired Excel file
            logger.info(f"Creating repaired file: {output_file}")
            with zipfile.ZipFile(output_file, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
                for root, _, files in os.walk(temp_dir):
                    for file in files:
                        file_path = Path(root) / file
                        arcname = file_path.relative_to(temp_dir)
                        zip_ref.write(file_path, arcname)

            logger.info(f"✓ Excel file repaired successfully: {output_file}")
            return output_file
        else:
            logger.info(f"No corruption found - using original file")
            return input_file

    except Exception as e:
        logger.error(f"Error during repair attempt: {e}")
        logger.info(f"Falling back to original file")
        return input_file

    finally:
        # Clean up temporary directory
        if temp_dir.exists():
            shutil.rmtree(temp_dir)


def _replace_external_formulas_with_text(sheet_file: Path, logger: logging.Logger) -> bool:
    """
    Replace cells containing external formulas and linked data types with the text '#REF'.

    Args:
        sheet_file: Path to the worksheet XML file
        logger: Logger instance

    Returns:
        bool: True if any modifications were made, False otherwise
    """
    try:
        # Parse the XML with namespace handling
        ET.register_namespace('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')
        ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')

        tree = ET.parse(sheet_file)
        root = tree.getroot()

        # Define namespace
        ns = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

        modified = False
        external_ref_count = 0
        linked_data_count = 0

        # Linked data type indicators
        linked_data_functions = ['FIELDVALUE', 'FILTERXML', 'ANCHORARRAY', '_xlfn.FIELDVALUE']

        # Find all cell elements
        for cell in root.findall('.//ss:c', ns):
            should_replace = False

            # Check if cell has a formula
            formula = cell.find('ss:f', ns)

            if formula is not None:
                formula_text = formula.text or ''

                # Check if formula references external workbook (contains '[' and ']')
                if '[' in formula_text and ']' in formula_text:
                    should_replace = True
                    external_ref_count += 1

                # Check if formula is a linked data type formula
                elif any(func in formula_text.upper() for func in linked_data_functions):
                    should_replace = True
                    linked_data_count += 1

            # Check for error type cells
            cell_type_attr = cell.get('t')
            if cell_type_attr == 'e':
                v_elem = cell.find('ss:v', ns)
                if v_elem is not None and v_elem.text in ['#REF!', '#VALUE!', '#N/A']:
                    should_replace = True

            if should_replace:
                # Remove the formula element if present
                if formula is not None:
                    cell.remove(formula)

                # Set cell type to inline string (str)
                cell.set('t', 'inlineStr')

                # Remove any existing value element
                for v in cell.findall('ss:v', ns):
                    cell.remove(v)

                # Remove any existing inline string
                for is_elem in cell.findall('ss:is', ns):
                    cell.remove(is_elem)

                # Add inline string value
                is_elem = ET.SubElement(cell, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}is')
                t_elem = ET.SubElement(is_elem, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                t_elem.text = '#REF'

                modified = True

        if modified:
            # Write back to file
            tree.write(sheet_file, encoding='utf-8', xml_declaration=True)
            details = []
            if external_ref_count > 0:
                details.append(f"{external_ref_count} external refs")
            if linked_data_count > 0:
                details.append(f"{linked_data_count} linked data")
            logger.info(f"  ✓ Repaired {sheet_file.name} ({', '.join(details)})")

        return modified

    except Exception as e:
        logger.warning(f"  ⚠ Could not process {sheet_file.name}: {e}")
        return False


def safe_load_workbook(filename: str, logger: Optional[logging.Logger] = None, **kwargs):
    """
    Safely load an Excel workbook, automatically repairing it if corruption is detected.

    Args:
        filename: Path to the Excel file
        logger: Optional logger instance
        **kwargs: Additional arguments to pass to openpyxl.load_workbook

    Returns:
        Workbook object from openpyxl

    Raises:
        Exception: If the file cannot be loaded even after repair attempt
    """
    if logger is None:
        logger = logging.getLogger(__name__)

    try:
        # Try to load normally first
        return load_workbook(filename, **kwargs)
    except Exception as e:
        error_msg = str(e).lower()
        # Check if it's a corruption-related error
        if any(keyword in error_msg for keyword in ['external', 'link', 'corrupt', 'repair', 'damaged']):
            logger.warning(f"Excel file appears corrupted: {e}")
            logger.info("Attempting automatic repair...")

            # Attempt repair
            repaired_file = repair_excel_file(filename, logger=logger)

            # Try to load the repaired file
            try:
                wb = load_workbook(repaired_file, **kwargs)
                logger.info("✓ Successfully loaded repaired file")
                return wb
            except Exception as repair_error:
                logger.error(f"Failed to load even after repair: {repair_error}")
                raise
        else:
            # Not a corruption error, re-raise
            raise
